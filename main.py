import os
import json
import csv
from datetime import datetime

from kivy.app import App
from kivy.lang import Builder
from kivy.properties import DictProperty, ListProperty, StringProperty, NumericProperty
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from kivy.uix.spinner import SpinnerOption
from kivy.resources import resource_find

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# --------------------------
# Helpers para APK (Android)
# --------------------------
def asset_path(filename: str) -> str:
    """
    Encuentra archivos incluidos en la app (csv, png, etc).
    Funciona en PC y en APK.
    """
    p = resource_find(filename)
    if p:
        return p
    return os.path.join(os.path.dirname(__file__), filename)

def stamp_name(prefix: str, ext: str) -> str:
    return f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"

def fmt_miles(n: int) -> str:
    return f"{n:,}".replace(",", ".")

def make_base_label(codigo: str, marca: str, nombre: str) -> str:
    return f"{codigo} | {marca} | {nombre}"


# Packs por marca (tus valores)
PACKS = {
    "SAMSON":   {"rollo": 2700,  "caja": 86400},
    "SINCLAIR": {"rollo": 18000, "caja": 126000},
}


KV = r"""
#:import dp kivy.metrics.dp

<ColoredSpinnerOption@SpinnerOption>:
    markup: True
    halign: "left"
    valign: "middle"
    text_size: self.size
    padding: dp(10), dp(6)

BoxLayout:
    orientation: "vertical"
    padding: dp(12)
    spacing: dp(10)

    Label:
        text: "Contador PLU / Sticker"
        font_size: "20sp"
        size_hint_y: None
        height: self.texture_size[1] + dp(6)

    BoxLayout:
        size_hint_y: None
        height: dp(44)
        spacing: dp(8)

        TextInput:
            id: inp_buscar
            hint_text: "Buscar (ej: 4040, HUASO, LUCILA...)"
            multiline: False
            on_text: app.filtrar_lista(self.text)

        Spinner:
            id: sp_plu
            text: app.seleccion_display if app.seleccion_display else "Selecciona PLU"
            values: app.display_lista_filtrada
            on_text: app.on_select_display(self.text)
            markup: True
            option_cls: "ColoredSpinnerOption"

    Label:
        text: app.pack_info
        size_hint_y: None
        height: dp(22)

    BoxLayout:
        size_hint_y: None
        height: dp(44)
        spacing: dp(8)

        TextInput:
            id: inp_cantidad
            hint_text: "Cantidad (ej: 15.000)"
            multiline: False
            input_type: "number"
            on_text: app.formatear_miles(self)

        Button:
            text: "+1"
            size_hint_x: None
            width: dp(60)
            on_release: app.aplicar_delta(1)

        Button:
            text: "+100"
            size_hint_x: None
            width: dp(70)
            on_release: app.aplicar_delta(100)

        Button:
            text: app.btn_rollo_text
            on_release: app.aplicar_delta(int(app.btn_rollo_val))

        Button:
            text: app.btn_caja_text
            on_release: app.aplicar_delta(int(app.btn_caja_val))

    BoxLayout:
        size_hint_y: None
        height: dp(44)
        spacing: dp(8)

        Button:
            text: app.btn_cuarto_caja_text
            on_release: app.aplicar_delta(int(app.btn_cuarto_caja_val))

        Button:
            text: app.btn_media_caja_text
            on_release: app.aplicar_delta(int(app.btn_media_caja_val))

        Button:
            text: "Caja vacía (0)"
            on_release: app.marcar_caja_vacia()

        Button:
            text: "Sumar (cantidad)"
            on_release: app.sumar_desde_input()

        Button:
            text: "Restar (cantidad)"
            on_release: app.restar_desde_input()

    BoxLayout:
        size_hint_y: None
        height: dp(44)
        spacing: dp(8)

        Button:
            text: "Deshacer"
            on_release: app.deshacer()

        Button:
            text: "Reset todo"
            on_release: app.reset_todo()

        Button:
            text: "Exportar XLSX"
            on_release: app.exportar_xlsx()

        Button:
            text: "Exportar CSV"
            on_release: app.exportar_csv()

    Label:
        text: app.estado
        size_hint_y: None
        height: dp(20)

    Label:
        text: "Totales por PLU:"
        bold: True
        size_hint_y: None
        height: dp(24)

    ScrollView:
        do_scroll_x: False
        GridLayout:
            id: grid_totales
            cols: 1
            size_hint_y: None
            height: self.minimum_height
            spacing: dp(6)

    BoxLayout:
        size_hint_y: None
        height: dp(44)

        Label:
            text: "TOTAL GENERAL: [b]{}[/b]   |   TIPOS REVISADOS: [b]{}[/b]".format(app.total_general, app.total_tipos)
            markup: True
            font_size: "18sp"
"""


def cargar_catalogo_csv(path_csv: str):
    items = []
    if not path_csv or not os.path.exists(path_csv):
        return items

    with open(path_csv, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            codigo = (row.get("codigo") or "").strip()
            marca = (row.get("marca") or "").strip().upper()
            nombre = (row.get("nombre") or "").strip()
            if codigo and nombre and marca:
                items.append({"codigo": codigo, "marca": marca, "nombre": nombre})
    return items


class ContadorPLUApp(App):
    conteos = DictProperty({})
    revisados = DictProperty({})
    historial = ListProperty([])

    base_lista_completa = ListProperty([])
    base_lista_filtrada = ListProperty([])
    display_lista_filtrada = ListProperty([])

    seleccion_base = StringProperty("")
    seleccion_display = StringProperty("")

    estado = StringProperty("")
    total_general = StringProperty("0")
    total_tipos = StringProperty("0")

    pack_info = StringProperty("Selecciona un PLU para ver packs rápidos.")

    btn_rollo_text = StringProperty("+Rollo")
    btn_caja_text = StringProperty("+Caja")
    btn_rollo_val = NumericProperty(0)
    btn_caja_val = NumericProperty(0)

    btn_cuarto_caja_text = StringProperty("+¼ Caja")
    btn_media_caja_text = StringProperty("+½ Caja")
    btn_cuarto_caja_val = NumericProperty(0)
    btn_media_caja_val = NumericProperty(0)

    def build(self):
        self.title = "Contador PLU"
        root = Builder.load_string(KV)
        self._root = root

        # ✅ Carga de CSV compatible con APK
        csv_catalogo = asset_path("plu_catalogo.csv")
        catalog_items = cargar_catalogo_csv(csv_catalogo)

        self.catalog_map = {}
        base_labels = []
        for it in catalog_items:
            base = make_base_label(it["codigo"], it["marca"], it["nombre"])
            self.catalog_map[base] = it
            base_labels.append(base)

        self.base_lista_completa = base_labels
        self.base_lista_filtrada = base_labels[:]

        if not base_labels:
            self.estado = "No encontré plu_catalogo.csv o está vacío (o falta 'marca')."

        self._cargar_autosave()
        self._reconstruir_display_list()
        self.refrescar_totales()
        return root

    def _popup(self, msg, titulo="Aviso"):
        Popup(title=titulo, content=Label(text=msg), size_hint=(0.9, 0.35)).open()

    def _display_for_base(self, base_label: str) -> str:
        cnt = int(self.conteos.get(base_label, 0))
        revisado = bool(self.revisados.get(base_label, False))
        texto = f"{base_label}  ({fmt_miles(cnt)})"
        return f"[color=2ecc71]{texto}[/color]" if revisado else f"[color=7f8c8d]{texto}[/color]"

    def _strip_markup(self, s: str) -> str:
        return (s.replace("[color=2ecc71]", "")
                 .replace("[color=7f8c8d]", "")
                 .replace("[/color]", "")
                 .strip())

    def _to_base_from_display(self, display_text: str) -> str:
        t = self._strip_markup(display_text)
        if t.endswith(")"):
            pos = t.rfind("  (")
            if pos != -1:
                t = t[:pos].strip()
        return t

    def _reconstruir_display_list(self):
        self.display_to_base = {}
        display = []
        for base in self.base_lista_filtrada:
            d = self._display_for_base(base)
            self.display_to_base[d] = base
            display.append(d)
        self.display_lista_filtrada = display

        if self.seleccion_base:
            self.seleccion_display = self._display_for_base(self.seleccion_base)

    def filtrar_lista(self, texto: str):
        t = (texto or "").strip().lower()
        if not t:
            self.base_lista_filtrada = self.base_lista_completa[:]
        else:
            self.base_lista_filtrada = [x for x in self.base_lista_completa if t in x.lower()]
        self._reconstruir_display_list()

    def on_select_display(self, display_text: str):
        base = self.display_to_base.get(display_text) or self._to_base_from_display(display_text)
        if base not in self.catalog_map:
            return
        self.seleccion_base = base
        self.seleccion_display = self._display_for_base(base)
        self._actualizar_packs()

    def _key(self):
        k = (self.seleccion_base or "").strip()
        return k if k else None

    def formatear_miles(self, widget):
        if getattr(self, "_formateando", False):
            return
        texto = widget.text or ""
        digitos = "".join(ch for ch in texto if ch.isdigit())

        self._formateando = True
        try:
            widget.text = "" if digitos == "" else fmt_miles(int(digitos))
            widget.cursor = (len(widget.text), 0)
        finally:
            self._formateando = False

    def _leer_cantidad(self):
        txt = self._root.ids.inp_cantidad.text.strip()
        if not txt:
            return None
        try:
            txt = txt.replace(".", "").replace(",", "")
            n = int(txt)
            return n if n > 0 else None
        except ValueError:
            return None

    def _limpiar_cantidad(self):
        self._root.ids.inp_cantidad.text = ""

    def _actualizar_packs(self):
        key = self._key()
        if not key:
            self.pack_info = "Selecciona un PLU para ver packs rápidos."
            self._deshabilitar_packs()
            return

        it = self.catalog_map.get(key)
        marca = (it.get("marca") if it else "").strip().upper()

        if marca not in PACKS:
            self.pack_info = "Marca inválida en CSV (debe ser SAMSON o SINCLAIR)."
            self._deshabilitar_packs()
            return

        rollo = PACKS[marca]["rollo"]
        caja = PACKS[marca]["caja"]
        cuarto = caja // 4
        media = caja // 2

        self.btn_rollo_val = rollo
        self.btn_caja_val = caja
        self.btn_cuarto_caja_val = cuarto
        self.btn_media_caja_val = media

        self.btn_rollo_text = f"+Rollo {fmt_miles(rollo)}"
        self.btn_caja_text = f"+Caja {fmt_miles(caja)}"
        self.btn_cuarto_caja_text = f"+¼ Caja {fmt_miles(cuarto)}"
        self.btn_media_caja_text = f"+½ Caja {fmt_miles(media)}"

        self.pack_info = f"Marca: {marca} | Rollo {fmt_miles(rollo)} | Caja {fmt_miles(caja)}"

    def _deshabilitar_packs(self):
        self.btn_rollo_val = 0
        self.btn_caja_val = 0
        self.btn_cuarto_caja_val = 0
        self.btn_media_caja_val = 0
        self.btn_rollo_text = "+Rollo"
        self.btn_caja_text = "+Caja"
        self.btn_cuarto_caja_text = "+¼ Caja"
        self.btn_media_caja_text = "+½ Caja"

    def _aplicar_mov(self, key: str, delta: int):
        actual = int(self.conteos.get(key, 0))
        nuevo = actual + int(delta)
        if nuevo < 0:
            return False

        self.revisados[key] = True
        self.conteos[key] = nuevo

        self.historial.append({
            "key": key,
            "delta": int(delta),
            "ts": datetime.now().isoformat(timespec="seconds")
        })
        return True

    def aplicar_delta(self, delta_pos: int):
        key = self._key()
        if not key:
            self._popup("Selecciona un PLU/sticker.")
            return
        if int(delta_pos) <= 0:
            self._popup("Botón deshabilitado (revisa marca/packs).")
            return

        ok = self._aplicar_mov(key, int(delta_pos))
        if not ok:
            self._popup("No se puede dejar en negativo.")
            return

        self.estado = f"Sumado +{fmt_miles(int(delta_pos))}"
        self.refrescar_totales()
        self._autosave()
        self._reconstruir_display_list()

    def sumar_desde_input(self):
        key = self._key()
        if not key:
            self._popup("Selecciona un PLU/sticker.")
            return

        cant = self._leer_cantidad()
        if cant is None:
            self._popup("Escribe una cantidad válida. Ej: 15.000")
            return

        self._aplicar_mov(key, cant)
        self._limpiar_cantidad()
        self.estado = f"Sumado +{fmt_miles(cant)}"
        self.refrescar_totales()
        self._autosave()
        self._reconstruir_display_list()

    def restar_desde_input(self):
        key = self._key()
        if not key:
            self._popup("Selecciona un PLU/sticker.")
            return

        cant = self._leer_cantidad()
        if cant is None:
            self._popup("Escribe una cantidad válida. Ej: 15.000")
            return

        ok = self._aplicar_mov(key, -cant)
        if not ok:
            self._popup("No se puede dejar el conteo en negativo.")
            return

        self._limpiar_cantidad()
        self.estado = f"Restado -{fmt_miles(cant)}"
        self.refrescar_totales()
        self._autosave()
        self._reconstruir_display_list()

    def marcar_caja_vacia(self):
        key = self._key()
        if not key:
            self._popup("Selecciona un PLU/sticker.")
            return

        self.revisados[key] = True
        self.conteos[key] = 0
        self.historial.append({
            "key": key, "delta": 0,
            "ts": datetime.now().isoformat(timespec="seconds"),
            "op": "caja_vacia"
        })

        self.estado = "Marcado como caja vacía (0)"
        self.refrescar_totales()
        self._autosave()
        self._reconstruir_display_list()

    def deshacer(self):
        if not self.historial:
            self._popup("No hay acciones para deshacer.")
            return

        ult = self.historial.pop()
        key = ult["key"]
        delta = int(ult.get("delta", 0))
        op = ult.get("op", "")

        if op == "caja_vacia":
            self.conteos.pop(key, None)
            self.revisados.pop(key, None)
        else:
            actual = int(self.conteos.get(key, 0))
            revertido = actual - delta
            if revertido < 0:
                revertido = 0
            self.conteos[key] = revertido

        self.estado = "Deshecho"
        self.refrescar_totales()
        self._autosave()
        self._reconstruir_display_list()

    def reset_todo(self):
        self.conteos = {}
        self.revisados = {}
        self.historial = []
        self.estado = "Reset completo"
        self.refrescar_totales()
        self._autosave()
        self._reconstruir_display_list()

    def refrescar_totales(self):
        grid = self._root.ids.grid_totales
        grid.clear_widgets()

        keys = sorted(self.revisados.keys())
        total = 0

        for key in keys:
            cnt = int(self.conteos.get(key, 0))
            total += cnt
            grid.add_widget(Label(text=f"{key}  -  {fmt_miles(cnt)}", size_hint_y=None, height=26))

        self.total_general = fmt_miles(total)
        self.total_tipos = str(len(keys))

    def _ruta_autosave(self):
        return os.path.join(self.user_data_dir, "autosave_conteo.json")

    def _autosave(self):
        data = {
            "conteos": self.conteos,
            "revisados": self.revisados,
            "historial": self.historial[-2000:],
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
        with open(self._ruta_autosave(), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _cargar_autosave(self):
        ruta = self._ruta_autosave()
        if not os.path.exists(ruta):
            return
        try:
            with open(ruta, "r", encoding="utf-8") as f:
                data = json.load(f)

            self.conteos = data.get("conteos", {}) or {}
            self.revisados = data.get("revisados", {}) or {}

            if not self.revisados and self.conteos:
                self.revisados = {k: True for k in self.conteos.keys()}

            self.historial = data.get("historial", []) or []
            self.estado = f"Cargado autosave ({data.get('updated_at','')})"
        except Exception as e:
            self.estado = f"No se pudo cargar autosave: {e}"

    def _parse_label(self, base_label: str):
        parts = base_label.split(" | ", 2)
        codigo = parts[0].strip() if len(parts) > 0 else ""
        marca = parts[1].strip() if len(parts) > 1 else ""
        nombre = parts[2].strip() if len(parts) > 2 else ""
        return codigo, marca, nombre

    def exportar_csv(self):
        if not self.revisados:
            self._popup("No hay datos para exportar.")
            return

        nombre = stamp_name("conteo_plu", "csv")
        ruta = os.path.join(self.user_data_dir, nombre)

        with open(ruta, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Codigo", "Marca", "Nombre", "Cantidad", "Revisado"])
            for key in sorted(self.revisados.keys()):
                cnt = int(self.conteos.get(key, 0))
                codigo, marca, nombre_mat = self._parse_label(key)
                w.writerow([codigo, marca, nombre_mat, cnt, "SI"])

        self._popup(f"CSV guardado en:\n{ruta}", "Exportado ✅")

    def exportar_xlsx(self):
        if not self.revisados:
            self._popup("No hay datos para exportar.")
            return

        nombre = stamp_name("conteo_plu", "xlsx")
        ruta = os.path.join(self.user_data_dir, nombre)

        wb = Workbook()
        ws = wb.active
        ws.title = "Conteo PLU"

        headers = ["Codigo", "Marca", "Nombre", "Cantidad", "Revisado"]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="E6E6E6")
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True)
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")

        row = 2
        total = 0
        for key in sorted(self.revisados.keys()):
            cnt = int(self.conteos.get(key, 0))
            codigo, marca, nombre_mat = self._parse_label(key)

            ws.cell(row=row, column=1, value=codigo)
            ws.cell(row=row, column=2, value=marca)
            ws.cell(row=row, column=3, value=nombre_mat)
            ws.cell(row=row, column=4, value=cnt)
            ws.cell(row=row, column=5, value="SI")

            total += cnt
            row += 1

        ws.cell(row=row + 1, column=3, value="TOTAL GENERAL").font = Font(bold=True)
        ws.cell(row=row + 1, column=4, value=total).font = Font(bold=True)

        widths = [18, 12, 60, 14, 12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(ruta)
        self._popup(f"XLSX guardado en:\n{ruta}", "Exportado ✅")


if __name__ == "__main__":
    ContadorPLUApp().run()
